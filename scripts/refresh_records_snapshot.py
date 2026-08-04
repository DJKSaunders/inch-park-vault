#!/usr/bin/env python3
"""Replace one season in the compact Vault records from cumulative snapshots.

The supplied batting and bowling XML files are complete season snapshots, not
append-only deltas.  The averages workbook is likewise an all-time snapshot,
so its boundary totals replace the existing totals.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    from scripts.records_quality import apply_administrative_no_play_rule
except ModuleNotFoundError:
    from records_quality import apply_administrative_no_play_rule


ROOT = Path(__file__).resolve().parents[1]


def normalize_player_name(candidate: str | None) -> str:
    if not candidate:
        return ""
    return re.sub(
        r"\s+\((?:SM'20|SM)\)\s*$", "", str(candidate), flags=re.IGNORECASE
    ).strip()


def normalize_team(candidate: str) -> str:
    if candidate == "Women's Premier (SM combined)":
        return "Women's"
    return candidate


def xml_value(node: ET.Element, field: str) -> str:
    return (node.findtext(field) or "").strip()


def xml_number(node: ET.Element, field: str) -> int:
    candidate = xml_value(node, field)
    try:
        return int(candidate)
    except ValueError:
        return 0


def xml_boolean(node: ET.Element, field: str) -> bool:
    return xml_value(node, field).casefold() == "true"


def xml_player_name(node: ET.Element) -> str:
    return normalize_player_name(
        " ".join(
            f"{xml_value(node, 'FirstName')} {xml_value(node, 'Surname')}".split()
        )
    )


def appearance_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row["player"].casefold(),
        row["date"],
        row["team"],
        row["matchType"],
        row["opposition"],
    )


def parse_batting_snapshot(path: Path, season: int) -> tuple[list[list[Any]], dict]:
    root = ET.parse(path).getroot()
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for node in root.findall("BattingPerfomance"):
        record_date = xml_value(node, "FixDate")[:10]
        if not record_date or int(record_date[:4]) != season:
            continue
        score = xml_value(node, "Score")
        did_not_bat = score.casefold() == "dnb"
        row = {
            "player": xml_player_name(node),
            "season": season,
            "team": normalize_team(xml_value(node, "TeamName")),
            "matchType": xml_value(node, "Type_Desc"),
            "opposition": xml_value(node, "Opposition"),
            "date": record_date,
            "runs": None if did_not_bat else int(score),
            "notOut": xml_boolean(node, "notout"),
            "didNotBat": did_not_bat,
            "catches": xml_number(node, "catches"),
            "stumpings": xml_number(node, "stumpings"),
            "runOuts": xml_number(node, "runouts"),
        }
        grouped[appearance_key(row)].append(row)

    output: list[list[Any]] = []
    suppressed_dnb = []
    collapsed_dnb = []
    multiple_innings = []
    for key in sorted(grouped):
        rows = grouped[key]
        innings = [row for row in rows if not row["didNotBat"]]
        dnb_rows = [row for row in rows if row["didNotBat"]]
        fielding = {
            field: max((row[field] for row in rows), default=0)
            for field in ("catches", "stumpings", "runOuts")
        }
        if innings:
            retained = innings
            if dnb_rows:
                suppressed_dnb.append(
                    {
                        "player": innings[0]["player"],
                        "date": innings[0]["date"],
                        "team": innings[0]["team"],
                        "opposition": innings[0]["opposition"],
                        "suppressedRows": len(dnb_rows),
                    }
                )
            if len(innings) > 1:
                multiple_innings.append(
                    {
                        "player": innings[0]["player"],
                        "date": innings[0]["date"],
                        "team": innings[0]["team"],
                        "opposition": innings[0]["opposition"],
                        "innings": len(innings),
                    }
                )
        else:
            retained = [dnb_rows[0]]
            if len(dnb_rows) > 1:
                collapsed_dnb.append(
                    {
                        "player": dnb_rows[0]["player"],
                        "date": dnb_rows[0]["date"],
                        "team": dnb_rows[0]["team"],
                        "opposition": dnb_rows[0]["opposition"],
                        "collapsedRows": len(dnb_rows) - 1,
                    }
                )

        for position, row in enumerate(retained):
            row = dict(row)
            for field in fielding:
                row[field] = fielding[field] if position == 0 else 0
            output.append(
                [
                    row["player"],
                    row["season"],
                    row["team"],
                    row["matchType"],
                    row["opposition"],
                    row["date"],
                    row["runs"],
                    row["notOut"],
                    row["didNotBat"],
                    row["catches"],
                    row["stumpings"],
                    row["runOuts"],
                ]
            )

    return output, {
        "sourceRowCount": sum(len(rows) for rows in grouped.values()),
        "outputRowCount": len(output),
        "suppressedDnbRows": suppressed_dnb,
        "collapsedDuplicateDnbRows": collapsed_dnb,
        "multipleBattingInnings": multiple_innings,
    }


def parse_bowling_snapshot(path: Path, season: int) -> list[list[Any]]:
    output = []
    root = ET.parse(path).getroot()
    for node in root.findall("Fixture"):
        record_date = xml_value(node, "FixDate")[:10]
        if not record_date or int(record_date[:4]) != season:
            continue
        output.append(
            [
                xml_player_name(node),
                season,
                normalize_team(xml_value(node, "TeamName")),
                xml_value(node, "Type_Desc"),
                xml_value(node, "Opposition"),
                record_date,
                xml_number(node, "totalballs"),
                xml_number(node, "Maidens"),
                xml_number(node, "Runs"),
                xml_number(node, "Wickets"),
            ]
        )
    return output


def parse_all_time_boundaries(path: Path) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next(
        name for name in workbook.sheetnames if name.strip().casefold().startswith("bat ")
    )
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = next(rows)
    index = {str(header).casefold(): position for position, header in enumerate(headers)}
    boundaries: dict[str, list[Any]] = {}
    for row in rows:
        first_name = row[index["firstname"]] or ""
        surname = row[index["surname"]] or ""
        player = normalize_player_name(" ".join(f"{first_name} {surname}".split()))
        if not player:
            continue
        key = player.casefold()
        candidate = [
            player,
            row[index["fours"]] if isinstance(row[index["fours"]], (int, float)) else 0,
            row[index["sixes"]] if isinstance(row[index["sixes"]], (int, float)) else 0,
        ]
        existing = boundaries.get(key)
        if existing:
            existing[1] = max(existing[1], candidate[1])
            existing[2] = max(existing[2], candidate[2])
        else:
            boundaries[key] = candidate
    return sorted(boundaries.values(), key=lambda row: row[0].casefold())


def refresh_payload(
    payload: dict[str, Any],
    season: int,
    batting: list[list[Any]],
    bowling: list[list[Any]],
    boundaries: list[list[Any]],
) -> dict[str, Any]:
    payload["batting"] = [row for row in payload["batting"] if row[1] != season] + batting
    payload["bowling"] = [row for row in payload["bowling"] if row[1] != season] + bowling
    payload["boundaries"] = boundaries

    rows = payload["batting"] + payload["bowling"]
    dates = [row[5] for row in rows if row[5]]
    seasons = sorted({row[1] for row in rows if row[1]})
    players = sorted({row[0] for row in rows if row[0]})
    meta = payload["meta"]
    meta.update(
        {
            "seasonStart": seasons[0],
            "seasonEnd": seasons[-1],
            "seasonCount": len(seasons),
            "recordCount": len(rows),
            "playerCount": len(players),
            "asOfDate": max(dates),
            "teams": sorted({row[2] for row in rows if row[2]}),
            "matchTypes": sorted({row[3] for row in rows if row[3]}),
            "oppositions": sorted({row[4] for row in rows if row[4]}),
            "playerNames": players,
        }
    )
    return payload


def atomic_json_write(path: Path, payload: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            if compact:
                json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
            else:
                json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
            handle.write("\n")
        os.replace(temporary_name, path)
    except Exception:
        try:
            os.unlink(temporary_name)
        except FileNotFoundError:
            pass
        raise


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("records", type=Path)
    parser.add_argument("batting_xml", type=Path)
    parser.add_argument("bowling_xml", type=Path)
    parser.add_argument("averages", type=Path)
    parser.add_argument("--season", type=int, required=True)
    parser.add_argument(
        "--scorecard-root", type=Path, default=ROOT / "public/data/scorecards"
    )
    parser.add_argument(
        "--quality-out", type=Path, default=ROOT / "public/data/records-quality.json"
    )
    args = parser.parse_args()

    payload = json.loads(args.records.read_text(encoding="utf-8"))
    batting, normalization = parse_batting_snapshot(args.batting_xml, args.season)
    bowling = parse_bowling_snapshot(args.bowling_xml, args.season)
    boundaries = parse_all_time_boundaries(args.averages)
    refresh_payload(payload, args.season, batting, bowling, boundaries)
    administrative = apply_administrative_no_play_rule(payload, args.scorecard_root)
    payload["meta"]["seasonSnapshot"] = {
        "season": args.season,
        "battingSource": args.batting_xml.name,
        "bowlingSource": args.bowling_xml.name,
        "averagesSource": args.averages.name,
    }
    quality = {
        "season": args.season,
        "normalization": normalization,
        "administrativeNoPlay": administrative,
    }
    atomic_json_write(args.records, payload, compact=True)
    atomic_json_write(args.quality_out, quality)
    print(
        json.dumps(
            {
                "records": str(args.records),
                "asOfDate": payload["meta"]["asOfDate"],
                "battingRows": len(payload["batting"]),
                "bowlingRows": len(payload["bowling"]),
                "boundaryPlayers": len(payload["boundaries"]),
                "suppressedDnbRows": sum(
                    row["suppressedRows"]
                    for row in normalization["suppressedDnbRows"]
                ),
                "collapsedDuplicateDnbRows": sum(
                    row["collapsedRows"]
                    for row in normalization["collapsedDuplicateDnbRows"]
                ),
                "suppressedAdministrativeRows": administrative["removedRowCount"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
