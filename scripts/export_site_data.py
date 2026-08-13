import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from records_quality import apply_administrative_no_play_rule


if len(sys.argv) not in (3, 6):
    raise SystemExit(
        "Usage: export_site_data.py <records.xlsx> <output.json> "
        "[<season-batting.xml> <season-bowling.xml> <season-averages.xlsx>]"
    )

workbook_path, output_path = sys.argv[1:3]
season_files = sys.argv[3:] if len(sys.argv) == 6 else None
workbook = load_workbook(workbook_path, read_only=True, data_only=True)
sheet = workbook["Unified Records"]
rows = sheet.iter_rows(values_only=True)
headers = next(rows)
index = {header: position for position, header in enumerate(headers)}


def value(row, column):
    return row[index[column]]


def number(row, column):
    candidate = value(row, column)
    return candidate if isinstance(candidate, (int, float)) else 0


def date_value(candidate):
    if isinstance(candidate, (datetime, date)):
        return candidate.isoformat()[:10]
    return candidate or ""


def normalize_player_name(candidate):
    if not candidate:
        return candidate
    value = re.sub(r"\s+\((?:SM'20|SM)\)\s*$", "", str(candidate), flags=re.IGNORECASE).strip()
    return {
        "kiran sv": "Kiran Sankaran",
        "srini m": "Srini Muthuraman",
    }.get(value.casefold(), value)


def normalize_team(candidate):
    if candidate == "Women's Premier (SM combined)":
        return "Women's"
    return candidate


def normalize_match_type(team, candidate):
    return "Friendly" if team == "Mitres" and candidate == "League" else candidate


batting = []
bowling = []
boundary_totals = {}
record_dates = []
players = set()
seasons = set()
teams = set()
match_types = set()
oppositions = set()

for row in rows:
    player = normalize_player_name(value(row, "Player Name"))
    season = number(row, "Season")
    team = normalize_team(value(row, "Team"))
    match_type = normalize_match_type(team, value(row, "Match Type"))
    opposition = value(row, "Opposition")
    record_date = date_value(value(row, "Date"))
    common = [
        player,
        season,
        team,
        match_type,
        opposition,
        record_date,
    ]

    if record_date:
        record_dates.append(record_date)
    players.add(player)
    seasons.add(season)
    teams.add(team)
    match_types.add(match_type)
    oppositions.add(opposition)

    if value(row, "Record Type") == "Batting":
        batting.append(
            common
            + [
                value(row, "Batting Runs"),
                bool(value(row, "Not Out")),
                bool(value(row, "Did Not Bat")),
                number(row, "Catches"),
                number(row, "Stumpings"),
                number(row, "Run Outs"),
            ]
        )
    elif value(row, "Record Type") == "Bowling":
        bowling.append(
            common
            + [
                number(row, "Balls Bowled"),
                number(row, "Maidens"),
                number(row, "Runs Conceded"),
                number(row, "Wickets"),
            ]
        )

summary_sheet = workbook["All-Time Batting"]
summary_rows = summary_sheet.iter_rows(values_only=True)
summary_headers = next(summary_rows)
summary_index = {
    header: position for position, header in enumerate(summary_headers)
}

for row in summary_rows:
    first_name = row[summary_index["First Name"]] or ""
    surname = row[summary_index["Surname"]] or ""
    player_name = normalize_player_name(" ".join(f"{first_name} {surname}".split()))
    fours = row[summary_index["Fours"]]
    sixes = row[summary_index["Sixes"]]
    if player_name:
        boundary_totals[player_name.casefold()] = [
            player_name,
            fours if isinstance(fours, (int, float)) else 0,
            sixes if isinstance(sixes, (int, float)) else 0,
        ]


def xml_value(node, field):
    return (node.findtext(field) or "").strip()


def xml_number(node, field):
    candidate = xml_value(node, field)
    try:
        return int(candidate)
    except ValueError:
        return 0


def xml_boolean(node, field):
    return xml_value(node, field).casefold() == "true"


def xml_player_name(node):
    return normalize_player_name(
        " ".join(
            f"{xml_value(node, 'FirstName')} {xml_value(node, 'Surname')}".split()
        )
    )


def add_common_metadata(player, season, team, match_type, opposition, record_date):
    players.add(player)
    seasons.add(season)
    teams.add(team)
    match_types.add(match_type)
    oppositions.add(opposition)
    if record_date:
        record_dates.append(record_date)


if season_files:
    batting_xml_path, bowling_xml_path, averages_path = season_files

    batting_root = ET.parse(batting_xml_path).getroot()
    for node in batting_root.findall("BattingPerfomance"):
        player = xml_player_name(node)
        record_date = xml_value(node, "FixDate")[:10]
        season = int(record_date[:4])
        team = normalize_team(xml_value(node, "TeamName"))
        match_type = normalize_match_type(team, xml_value(node, "Type_Desc"))
        opposition = xml_value(node, "Opposition")
        score = xml_value(node, "Score")
        did_not_bat = score.casefold() == "dnb"

        batting.append(
            [
                player,
                season,
                team,
                match_type,
                opposition,
                record_date,
                None if did_not_bat else int(score),
                xml_boolean(node, "notout"),
                did_not_bat,
                xml_number(node, "catches"),
                xml_number(node, "stumpings"),
                xml_number(node, "runouts"),
            ]
        )
        add_common_metadata(
            player, season, team, match_type, opposition, record_date
        )

    bowling_root = ET.parse(bowling_xml_path).getroot()
    for node in bowling_root.findall("Fixture"):
        player = xml_player_name(node)
        record_date = xml_value(node, "FixDate")[:10]
        season = int(record_date[:4])
        team = normalize_team(xml_value(node, "TeamName"))
        match_type = normalize_match_type(team, xml_value(node, "Type_Desc"))
        opposition = xml_value(node, "Opposition")

        bowling.append(
            [
                player,
                season,
                team,
                match_type,
                opposition,
                record_date,
                xml_number(node, "totalballs"),
                xml_number(node, "Maidens"),
                xml_number(node, "Runs"),
                xml_number(node, "Wickets"),
            ]
        )
        add_common_metadata(
            player, season, team, match_type, opposition, record_date
        )

    averages_workbook = load_workbook(
        averages_path, read_only=True, data_only=True
    )
    batting_sheet_name = next(
        name
        for name in averages_workbook.sheetnames
        if name.strip().casefold().startswith("bat ")
    )
    averages_sheet = averages_workbook[batting_sheet_name]
    averages_rows = averages_sheet.iter_rows(values_only=True)
    averages_headers = next(averages_rows)
    averages_index = {
        str(header).casefold(): position
        for position, header in enumerate(averages_headers)
    }

    for row in averages_rows:
        first_name = row[averages_index["firstname"]] or ""
        surname = row[averages_index["surname"]] or ""
        player_name = normalize_player_name(
            " ".join(f"{first_name} {surname}".split())
        )
        if not player_name:
            continue
        key = player_name.casefold()
        existing = boundary_totals.setdefault(key, [player_name, 0, 0])
        existing[1] += row[averages_index["fours"]] or 0
        existing[2] += row[averages_index["sixes"]] or 0

players = sorted(filter(None, players))
seasons = sorted(filter(None, seasons))
boundaries = sorted(boundary_totals.values(), key=lambda row: row[0].casefold())

payload = {
    "meta": {
        "title": "The Inch Park Vault",
        "seasonStart": seasons[0],
        "seasonEnd": seasons[-1],
        "recordCount": len(batting) + len(bowling),
        "playerCount": len(players),
        "seasonCount": len(seasons),
        "asOfDate": max(record_dates) if record_dates else "",
        "teams": sorted(filter(None, teams)),
        "matchTypes": sorted(filter(None, match_types)),
        "oppositions": sorted(filter(None, oppositions)),
        "playerNames": players,
        "generatedFrom": os.path.basename(workbook_path),
    },
    "batting": batting,
    "bowling": bowling,
    "boundaries": boundaries,
}

root = Path(__file__).resolve().parents[1]
scorecard_root = root / "public" / "data" / "scorecards"
quality = apply_administrative_no_play_rule(payload, scorecard_root)

os.makedirs(os.path.dirname(output_path), exist_ok=True)
with open(output_path, "w", encoding="utf-8") as handle:
    json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))

print(
    json.dumps(
        {
            "outputPath": output_path,
            "bytes": os.path.getsize(output_path),
            "battingRows": len(payload["batting"]),
            "bowlingRows": len(payload["bowling"]),
            "suppressedAdministrativeRows": quality["removedRowCount"],
        }
    )
)
